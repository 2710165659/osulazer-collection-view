from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .models import BeatmapEntry, CollectionInfo

MODE_EXPORT_ORDER = ("osu", "taiko", "ctb", "mania")


def _safe_sheet_name(base: str, used: set[str]) -> str:
    """Excel sheet 名长度有限，这里统一做清洗和去重。"""
    cleaned = (base or "Collection").replace("/", " ").replace("\\", " ").replace("*", " ").replace("?", " ")
    cleaned = cleaned.replace("[", "(").replace("]", ")").replace(":", " ").strip() or "Collection"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 1
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _safe_filename(base: str, suffix: str) -> str:
    """按 Windows 文件名规则做最小清洗，避免 zip 内文件名非法。"""
    cleaned = (base or "export").strip()
    for char in '<>:"/\\|?*':
        cleaned = cleaned.replace(char, "_")
    return f"{cleaned}{suffix}"


def _write_header_row(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(sheet, minimum: int = 10, maximum: int = 40) -> None:
    """按内容估算列宽，避免导出后肉眼难看。"""
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, minimum), maximum)


def export_current_view(path: Path, sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
    """导出当前表格所见即所得的谱面列表。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_name(sheet_name or "Beatmaps", set())
    _write_header_row(sheet, headers)

    for row in rows:
        sheet.append(row)

    _autosize_columns(sheet)
    workbook.save(path)


def build_mode_workbook(
    collections: list[CollectionInfo],
    mode: str,
    headers: list[str],
    row_builder,
) -> Workbook:
    """单个模式一个 Excel，Summary + 每个收藏夹一个 sheet。"""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_header_row(summary, ["Collection", "Mode", "Visible Items", "Missing Items", "Last Modified"])

    used_names = {"Summary"}
    for collection in collections:
        items = collection.items_for_mode(mode)
        if not items:
            continue

        summary.append(
            [
                collection.name,
                mode,
                len(items),
                sum(1 for item in items if item.missing),
                collection.last_modified_text,
            ]
        )

        sheet = workbook.create_sheet(_safe_sheet_name(collection.name, used_names))
        _write_header_row(sheet, headers)
        for item in items:
            sheet.append(row_builder(collection.name, item))
        _autosize_columns(sheet)

    _autosize_columns(summary, maximum=28)
    return workbook


def export_all_modes_zip(path: Path, collections: list[CollectionInfo], headers: list[str], row_builder) -> None:
    """导出四个模式的压缩包，每个模式一个 workbook。"""
    with ZipFile(path, "w", compression=ZIP_DEFLATED) as archive:
        for mode in MODE_EXPORT_ORDER:
            workbook = build_mode_workbook(collections, mode, headers, row_builder)
            buffer = BytesIO()
            workbook.save(buffer)
            workbook.close()
            archive.writestr(_safe_filename(f"collections_{mode}", ".xlsx"), buffer.getvalue())
